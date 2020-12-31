#Script to backtest trades

from yahoo_fin import stock_info as si
from datetime import datetime as dt
import openpyxl as op
import numpy as np

testing = False #override trading hours restriction  

def record(time, ticker, price, quantity): #if all checks pass, record the trade
    wb = op.load_workbook(filename = 'log.xlsx')
    
    sh = wb['Trades']
    prevrow = sh.max_row
    currRow = prevrow + 1
    
    sh['A' + str(currRow)] = time
    sh['B' + str(currRow)] = ticker
    sh['C' + str(currRow)] = price
    sh['D' + str(currRow)] = quantity
    sh['E' + str(currRow)] = np.round(sh['E' + str(prevrow)].value - sh['C' + str(currRow)].value*quantity, 2)
    
    sh = wb['Stocks'] #update stocks sheet
    nameFound = False

    for cell in sh['A']:
        if(cell.value == ticker): #if the stock already exists
            if quantity < 0:
                sh['B' + str(cell.row)] = time
            sh['C' + str(cell.row)] = sh['C' + str(cell.row)].value + quantity
            nameFound = True
    if nameFound == False: #if the stock does not exist (has never been bought before)
        currRow = sh.max_row + 1
        sh['A' + str(currRow)] = ticker
        sh['B' + str(currRow)] = time #this should only execute for an initial buy
        sh['C' + str(currRow)] = quantity
    wb.save(filename = 'log.xlsx')

def checkHours(t): #check trading hours
    wb = op.load_workbook(filename = 'log.xlsx')
        
    if (t.hour >= 9 and t.hour <= 16) or(t.hour == 16 and t.minute < 30) or testing == True:
        return True
    else:
        print('Outside of Trading Hours')
        sh = wb['Errors']
        sh['A1'] = 'YES'
        wb.save(filename = 'log.xlsx')
        return False

def checkBalance(price, quantity): #check balance
    wb = op.load_workbook(filename = 'log.xlsx')
    sh = wb['Trades']
    prevrow = sh.max_row
    
    if  price*quantity <= sh['E' + str(prevrow)].value:
        return True
    else:
        print('Insufficient Funds')
        sh = wb['Errors']
        sh['A2'] = 'YES'
        wb.save(filename = 'log.xlsx')
        return False
    
def checkQty(ticker, quantity): #check quantity of stock
    wb = op.load_workbook(filename = 'log.xlsx')
    sh = wb['Stocks'] #add stock to spreadsheet
    nameFound = False

    for cell in sh['A']: #look for the ticker
        if(cell.value == ticker):
            if sh['C' + str(cell.row)].value >= quantity: #see if there are enough stocks to sell
                return True
            else:
                print('Insufficient Stock Quantity')
                sh = wb['Errors']
                sh['A3'] = 'YES'
                wb.save(filename = 'log.xlsx')
                return False
            nameFound = True
    if nameFound == False: #if the stock isn't even on the list
        print('Insufficient Stock Quantity')
        sh = wb['Errors']
        sh['A3'] = 'YES'
        wb.save(filename = 'log.xlsx')
        return False

def checkDayTrades(ticker, time):
    wb = op.load_workbook(filename = 'log.xlsx')
    sh = wb['Stocks'] #add stock to spreadsheet
    
    for cell in sh['A']: #look for the ticker
        if(cell.value == ticker):
            if sh['B' + str(cell.row)].value.day == time.day:
                print('Day Trade Executed')
                sh = wb['Errors']
                sh['A4'] = sh['A4'].value + 1
                wb.save(filename = 'log.xlsx')
    
def qty(ticker):
    wb = op.load_workbook(filename = 'log.xlsx')
    sh = wb['Stocks']
    nameFound = False
    
    for cell in sh['A']:
        if(cell.value == ticker):
            nameFound = True
            return sh['C' + str(cell.row)].value
    if nameFound == False:
        return 0
    
def balance():
    wb = op.load_workbook(filename = 'log.xlsx')
    sh = wb['Trades']
    amount = sh['E' + str(sh.max_row)].value
    return amount

def refresh(): # THIS NEEDS TO BE FINISHED STILL
    t = dt.now()
    
    if (t.hour >= 9 and t.hour <= 16) or (t.hour == 16 and t.minute < 30) or testing == True: #try to execute a trade
        wb = op.load_workbook(filename = 'log.xlsx', read_only = True)
        sh = wb['Limit']
        for cell in sh['B']:
            if si.get_live_price(cell.value) <= sh['C' + str(cell.row)].value:
                buy(ticker, sh['D' + str(cell.row)])
                sh.delete_rows(cell.row)
                #DELETE ROWS
  # CASE FOR REMOVING LIMIT ORDERS AT THE END OF THE DAY

def buy(ticker, quantity):
    #check trading hours, balance
    
    t = dt.now()
    price = si.get_live_price(ticker)
    
    if checkHours(t): #check trading hours
        if checkBalance(price, quantity): #check balance
            record(t, ticker, price, quantity) #record the trade
            
def sell(ticker, quantity):
    #check trading hours, quantity of stocks to be sold, day trades
    
    t = dt.now()
    price = si.get_live_price(ticker)
    checkDayTrades(ticker, t)
    
    if checkHours(t): #check trading hours
        if checkQty(ticker, quantity): #check quantity of stocks
            record(t, ticker, price, -quantity) #record the trade
            
def limitOrder(ticker, quantity, maxPrice):
    wb = op.load_workbook(filename = 'log.xlsx')
    sh = wb['Limit']
    
    currRow = sh.max_row + 1
    
    sh['A' + str(currRow)] = dt.now()
    sh['B' + str(currRow)] = ticker
    sh['C' + str(currRow)] = maxPrice
    sh['D' + str(currRow)] = quantity
    
    wb.save(filename = 'log.xlsx')