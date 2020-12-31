import getStockData as sd
import time as tm
import numpy as np
import pandas as pd
import openpyxl as op
from statistics import mean

mData = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I'] #columns containing each type of data
bData = ['J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q']
wData = ['R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y']
aData = ['Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG']

# pd.set_option('display.max_rows', None, 'display.max_columns', None)

def populate(ticker, stockConstants, rowNumber, sh):
    for j in mData:
        sh[j + str(rowNumber)] = stockConstants.loc[mData.index(j),'m']
    for k in bData:
        sh[k + str(rowNumber)] = stockConstants.loc[bData.index(k),'b']
    for l in wData:
        sh[l + str(rowNumber)] = stockConstants.loc[wData.index(l),'w']
    for m in aData:
        sh[m + str(rowNumber)] = stockConstants.loc[aData.index(m),'a']

def filter(predictions):
    thresh = 0.1
    remList = []
    for i in range(len(predictions.index)):
        raw = predictions.iloc[i]
        ls = raw.loc['Pred (ls)']
        surf = raw.loc['Pred (surf)']
        avg = raw.loc['Average']
        if ((ls < 0 and surf > 0) or (ls > 0 and surf < 0)): #check if the signs agree
            remList.append(i)
        elif abs(avg) < thresh: #check if the avg is above the threshold
            remList.append(i)
    predictions.drop(predictions.index[remList], inplace = True)
    predictions.sort_values('Average', inplace = True, ascending = False)
    return predictions

def learn(tickerList, inter):
    wb = op.load_workbook(filename = 'stockData.xlsx') #open the workbook
    
    for i in tickerList: #for each of the tickers...
        print(i)
        stockConstants = sd.getStats(i, inter)
        
        sh = wb[inter]
        nameFound = False
        for cell in sh['A']: #loop through every cell in column 'A'    
            if(cell.value == i): #if the stock already exists...
                rowNumber = cell.row #the row number is the row with the ticker in it
                nameFound = True 
        if nameFound == False: #if the stock does not exist (has never been bought before)...
            rowNumber = sh.max_row + 1 #the row number is the next empty row
            sh['A' + str(rowNumber)] = i #enter the ticker name in column 'A'
        print('(Writing Data to Spreadsheet)\n')
        populate(i, stockConstants, rowNumber, sh) #populate the row with the data
        
        #save the doc
        wb.save(filename = 'stockData.xlsx') #save the workbook

def predict(tickerList, inter):
    wb = op.load_workbook(filename = 'stockData.xlsx')
    sh = wb[inter]
    surfPred = []
    leastSquaresPred = []
    
    for i in tickerList:
        tm.sleep(5)
        print('\n')
        print(i)
        x = sd.getCurrent(i, inter)
        
        nameFound = False
        
        for cell in sh['A']:
            if(cell.value == i): #if the stock already exists
                print('(Reading Data from Spreadsheet)')
                m = np.array([sh['B' + str(cell.row)].value,
                     sh['C' + str(cell.row)].value,
                     sh['D' + str(cell.row)].value,
                     sh['E' + str(cell.row)].value,
                     sh['F' + str(cell.row)].value,
                     sh['G' + str(cell.row)].value,
                     sh['H' + str(cell.row)].value,
                     sh['I' + str(cell.row)].value])
                b = np.array([sh['J' + str(cell.row)].value,
                     sh['K' + str(cell.row)].value,
                     sh['L' + str(cell.row)].value,
                     sh['M' + str(cell.row)].value,
                     sh['N' + str(cell.row)].value,
                     sh['O' + str(cell.row)].value,
                     sh['P' + str(cell.row)].value,
                     sh['Q' + str(cell.row)].value])
                w = np.array([sh['R' + str(cell.row)].value,
                     sh['S' + str(cell.row)].value,
                     sh['T' + str(cell.row)].value,
                     sh['U' + str(cell.row)].value,
                     sh['V' + str(cell.row)].value,
                     sh['W' + str(cell.row)].value,
                     sh['X' + str(cell.row)].value,
                     sh['Y' + str(cell.row)].value])
                a = np.array([sh['Z' + str(cell.row)].value,
                     sh['AA' + str(cell.row)].value,
                     sh['AB' + str(cell.row)].value,
                     sh['AC' + str(cell.row)].value,
                     sh['AD' + str(cell.row)].value,
                     sh['AE' + str(cell.row)].value,
                     sh['AF' + str(cell.row)].value,
                     sh['AG' + str(cell.row)].value])
                leastSquaresScore = (w*(m*x + b)).sum()
                surfScore = (a*x).sum()
                leastSquaresPred.append(leastSquaresScore)
                surfPred.append(surfScore)
    leastSquaresPred = np.round(leastSquaresPred, 4)
    surfPred = np.round(surfPred, 4)
    avg = (leastSquaresPred + surfPred)/2
    allPred = pd.DataFrame({'ticker': tickerList,
                            'Pred (ls)': leastSquaresPred,
                            'Pred (surf)': surfPred,
                            'Average': avg})
    return(allPred)
    
tickerList = ['RIO', 'CBAT', 'CNI', 'ANTM', 'CB', 'TFC', 'SO', 'DE', 'PBR', 'USB', 'BP', 'DUK', 'CI', 'FDX', 'BDX', 'CL', 'ZTS', 'SNAP', 'PLD', 'SNP', 'BEKE', 'INFY', 'NVDA', 'NIO', 'CCI', 'RIDE', 'ZM', 'BLNK', 'FUV', 'PLUG', 'SPGI', 'TJX']
print(len(tickerList))
inter = 'daily'
print('Estimated Time: %d minutes' %(np.size(tickerList)))
learn(tickerList, inter)
ranking = predict(tickerList, inter)
ranking = filter(ranking)
print(ranking)